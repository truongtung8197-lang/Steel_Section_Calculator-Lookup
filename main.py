import logging
import math
import os
import sys
from dataclasses import dataclass
from typing import Callable, Dict, List, Tuple

import openpyxl
from PySide6.QtCore import Qt
from PySide6.QtGui import QPixmap, QDoubleValidator
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QFormLayout,
    QGroupBox,
    QGridLayout,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QPushButton,
    QSizePolicy,
    QSpacerItem,
    QVBoxLayout,
    QWidget,
    QSplitter,
    QTabWidget,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QMenuBar,
    QMenu,
)

DENSITY_FACTOR = 7.85e-6  # kg per mm³

# Unit conversion factors (to mm)
UNIT_CONVERSION = {"mm": 1.0, "cm": 10.0, "m": 1000.0, "inch": 25.4}

# Đường dẫn sẽ tự động đi theo file .exe
# Tự động lấy đường dẫn của thư mục chứa file .py hoặc file .exe đang chạy
if getattr(sys, "frozen", False):
    # Nếu đang chạy từ file .exe đã đóng gói
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # Nếu đang chạy từ source code .py bằng phần mềm lập trình
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# Setup logging (after BASE_DIR is defined)
logging.basicConfig(
    filename=os.path.join(BASE_DIR, "app.log"),
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# Đường dẫn sẽ tự động đi theo file .exe
PNG_DIR = os.path.join(BASE_DIR, "STEEL TYPE png")
EXCEL_PATH = os.path.join(BASE_DIR, "alias.xlsx")
JSON_PATH = os.path.join(BASE_DIR, "steel_db.json")


@dataclass(frozen=True)
class SteelType:
    key: str
    name: str
    fields: List[Tuple[str, str]]
    calc: Callable[[dict], float]
    validator: Callable[[dict], None]
    image_file: str
    tooltip: str = ""


# --- Logic tính toán hình học (Tab 1) ---
def area_plate(v):
    return v["Length"] * v["Width"] * v["Thickness"]


def check_plate(v):
    if v["Length"] <= 0 or v["Width"] <= 0 or v["Thickness"] <= 0:
        raise ValueError()


def area_ishape(v):
    h, b, tw, tf = v["H"], v["B"], v["Tw"], v["Tf"]
    r = v.get("r1", 0)
    base_area = 2 * b * tf + (h - 2 * tf) * tw
    corner_area = (math.pi - 2) * r ** 2 if r > 0 else 0
    return base_area + corner_area


def check_ishape(v):
    r = v.get("r1", 0)
    if v["Tw"] >= v["B"] or (2 * v["Tf"]) >= v["H"]:
        raise ValueError()
    if r < 0 or r > min(v["Tw"], v["Tf"]) / 2:
        raise ValueError()


def area_channel(v):
    h, b, tw, tf = v["H"], v["B"], v["Tw"], v["Tf"]
    r = v.get("r1", 0)
    base_area = 2 * b * tf + (h - 2 * tf) * tw
    corner_area = 2 * (math.pi - 2) * r ** 2 if r > 0 else 0
    return base_area + corner_area


def check_channel(v):
    r = v.get("r1", 0)
    if v["Tw"] >= v["B"] or (2 * v["Tf"]) >= v["H"]:
        raise ValueError()
    if r < 0 or r > min(v["Tw"], v["Tf"]) / 2:
        raise ValueError()


def area_tsection(v):
    h, b, tw, tf = v["H"], v["B"], v["Tw"], v["Tf"]
    r = v.get("r1", 0)
    base_area = b * tf + (h - tf) * tw
    corner_area = 2 * (math.pi - 2) * r ** 2 if r > 0 else 0
    return base_area + corner_area


def check_tsection(v):
    r = v.get("r1", 0)
    if v["Tw"] >= v["B"] or v["Tf"] >= v["H"]:
        raise ValueError()
    if r < 0 or r > min(v["Tw"], v["Tf"]) / 2:
        raise ValueError()


def area_angle(v):
    a, b, t = v["Leg A"], v["Leg B"], v["Thickness"]
    r = v.get("r1", 0)
    base_area = t * (a + b - t)
    corner_area = (math.pi / 4 - 0.5) * r ** 2 if r > 0 else 0
    return base_area + corner_area


def check_angle(v):
    r = v.get("r1", 0)
    if v["Thickness"] >= v["Leg A"] or v["Thickness"] >= v["Leg B"]:
        raise ValueError()
    if r < 0 or r > v["Thickness"]:
        raise ValueError()


def area_rhs_shs(v):
    w, h, t = v["Width"], v["Height"], v["Thickness"]
    r = v.get("r1", 0)
    base_area = w * h - (w - 2 * t) * (h - 2 * t)
    if r > 0:
        ro = r
        ri = r - t
        corner_area = (4 - math.pi) * (ro ** 2 - ri ** 2)
        return base_area - corner_area
    return base_area


def check_rhs_shs(v):
    r = v.get("r1", 0)
    if (2 * v["Thickness"]) >= v["Width"] or (2 * v["Thickness"]) >= v["Height"]:
        raise ValueError()
    if r < 0 or r > v["Thickness"]:
        raise ValueError()


def area_chs(v):
    od, t = v["OD"], v["Thickness"]
    return math.pi / 4.0 * (od**2 - (od - 2 * t) ** 2)


def check_chs(v):
    if (2 * v["Thickness"]) >= v["OD"]:
        raise ValueError()


def area_rod(v):
    return math.pi * v["Diameter"] ** 2 / 4.0


def check_rod(v):
    if v["Diameter"] <= 0:
        raise ValueError()


STEEL_TYPES: List[SteelType] = [
    SteelType(
        "plate",
        "Plate",
        [("Length", "mm"), ("Width", "mm"), ("Thickness", "mm")],
        lambda v: area_plate(v) * DENSITY_FACTOR,
        check_plate,
        "PL.PNG",
        tooltip="Tấm thép phẳng. Tính khối lượng = Dài × Rộng × Dày × Mật độ thép",
    ),
    SteelType(
        "ih",
        "I Beam / H Beam",
        [("H", "mm"), ("B", "mm"), ("Tw", "mm"), ("Tf", "mm"), ("Length", "m")],
        lambda v: area_ishape(v) * v["Length"] * 1000.0 * DENSITY_FACTOR,
        check_ishape,
        "I.PNG",
        tooltip="Dầm I/H: H=chiều cao, B=chiều rộng cánh, Tw=dày nhịp, Tf=dày cánh",
    ),
    SteelType(
        "channel",
        "PFC / U Channel",
        [("H", "mm"), ("B", "mm"), ("Tw", "mm"), ("Tf", "mm"), ("Length", "m")],
        lambda v: area_channel(v) * v["Length"] * 1000.0 * DENSITY_FACTOR,
        check_channel,
        "U.PNG",
        tooltip="Thép hình U: H=chiều cao, B=chiều rộng đáy, Tw=dày nhịp, Tf=dày đáy",
    ),
    SteelType(
        "angle",
        "Angle / L Section",
        [("Leg A", "mm"), ("Leg B", "mm"), ("Thickness", "mm"), ("Length", "m")],
        lambda v: area_angle(v) * v["Length"] * 1000.0 * DENSITY_FACTOR,
        check_angle,
        "L.PNG",
        tooltip="Thép góc: Leg A/B=chiều dài 2 cạnh, Thickness=dày",
    ),
    SteelType(
        "rhs_shs",
        "RHS / SHS",
        [("Width", "mm"), ("Height", "mm"), ("Thickness", "mm"), ("Length", "m")],
        lambda v: area_rhs_shs(v) * v["Length"] * 1000.0 * DENSITY_FACTOR,
        check_rhs_shs,
        "RHS.PNG",
        tooltip="Thép hộp: Width=chiều rộng, Height=chiều cao, Thickness=dày thành",
    ),
    SteelType(
        "chs",
        "CHS / Pipe",
        [("OD", "mm"), ("Thickness", "mm"), ("Length", "m")],
        lambda v: area_chs(v) * v["Length"] * 1000.0 * DENSITY_FACTOR,
        check_chs,
        "CHS.PNG",
        tooltip="Thép ống: OD=đường kính ngoài, Thickness=dày thành",
    ),
    SteelType(
        "rod",
        "Rod / Round Bar",
        [("Diameter", "mm"), ("Length", "m")],
        lambda v: area_rod(v) * v["Length"] * 1000.0 * DENSITY_FACTOR,
        check_rod,
        "ROD.PNG",
        tooltip="Thép tròn: Diameter=đường kính",
    ),
    SteelType(
        "tsection",
        "T Section",
        [("H", "mm"), ("B", "mm"), ("Tw", "mm"), ("Tf", "mm"), ("Length", "m")],
        lambda v: area_tsection(v) * v["Length"] * 1000.0 * DENSITY_FACTOR,
        check_tsection,
        "T.PNG",
        tooltip="Thép chữ T: H=chiều cao, B=chiều rộng đỉnh, Tw=dày nhịp, Tf=dày đỉnh",
    ),
]


class ImageBox(QLabel):
    def __init__(self):
        super().__init__()
        self.setAlignment(Qt.AlignCenter)
        self.setMinimumSize(300, 300)
        self.setStyleSheet(
            "background: #ffffff; border: 1px solid #e2e8f0; border-radius: 12px; padding: 10px;"
        )
        self._current_path = None

    def set_image(self, path: str):
        self._current_path = path
        self._render()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._render()

    def _render(self):
        if not self._current_path or not os.path.exists(self._current_path):
            self.setText(
                "ℹ️ Technical Drawing Image Not Found\nPlace your PNGs in " + PNG_DIR
            )
            self.setPixmap(QPixmap())
            return
        pix = QPixmap(self._current_path)
        if pix.isNull():
            self.setText("⚠️ Failed to load image template.")
            self.setPixmap(QPixmap())
            return

        w = pix.size().width()
        h = pix.size().height()
        if w > 0 and h > 0:
            scaled = pix.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.setPixmap(scaled)
            self.setText("")


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Steel Management & Calculator Pro")
        self.resize(1150, 750)
        self.setMinimumSize(950, 650)

        self.excel_data: List[dict] = []
        self.current_raw_weight = None
        self.current_unit = "mm"  # Default unit
        self.load_data()

        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        self.create_menu_bar()
        self.init_calculator_tab()
        self.init_lookup_tab()
        self.apply_styles()

        logger.info("Application started successfully")

    def create_menu_bar(self):
        menubar = self.menuBar()

        # Help menu
        help_menu = menubar.addMenu("&Help")

        about_action = help_menu.addAction("&About")
        about_action.triggered.connect(self.show_about)

        help_action = help_menu.addAction("&User Guide")
        help_action.triggered.connect(self.show_help)

    def show_about(self):
        QMessageBox.about(
            self,
            "About Steel Management & Calculator Pro",
            "Steel Management & Calculator Pro v1.0\n\n"
            "A tool for calculating steel weight and looking up standard steel sections.\n\n"
            "Technology: Python 3.x, PySide6, openpyxl\n"
            "Developer: Development Team\n"
            "Year: 2024",
        )

    def show_help(self):
        help_text = """
        <h2>Steel Management & Calculator Pro - User Guide</h2>

        <h3>Tab 1: Manual Calculator</h3>
        <ol>
            <li><b>Select Steel Type:</b> Choose the type of steel section from the dropdown</li>
            <li><b>Enter Dimensions:</b> Input the section dimensions in the form fields</li>
            <li><b>Quantity:</b> Enter the quantity (default is 1)</li>
            <li><b>Result:</b> The calculated weight will appear automatically</li>
            <li><b>Copy:</b> Click "Copy Value" to copy the result to clipboard</li>
            <li><b>Clear:</b> Click "Clear" to reset all input fields</li>
        </ol>

        <h3>Tab 2: Steel Section Lookup</h3>
        <ol>
            <li><b>Select Steel Library:</b> Choose the steel library/database</li>
            <li><b>Search:</b> Enter keywords to search for sections (e.g., "WC 500" or "500")</li>
            <li><b>Select Profile:</b> Click on a profile from the list to view details</li>
            <li><b>Copy:</b> Click "Copy" buttons to copy individual values</li>
        </ol>

        <h3>Tips:</h3>
        <ul>
            <li>All dimensions are in millimeters (mm) unless otherwise specified</li>
            <li>Length for beams/channels is in meters (m)</li>
            <li>You can search with multiple keywords separated by spaces</li>
            <li>Hover over input fields for tooltips and guidance</li>
        </ul>

        <h3>Steel Density:</h3>
        <p>Standard steel density: 7850 kg/m³ (7.85e-6 kg/mm³)</p>
        """
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("User Guide")
        msg_box.setTextFormat(Qt.RichText)
        msg_box.setText(help_text)
        msg_box.exec()

    # ==========================================
    # LOGIC TAB 1: BỘ TÍNH TOÁN THỦ CÔNG
    # ==========================================
    def init_calculator_tab(self):
        calc_widget = QWidget()
        main_splitter = QSplitter(Qt.Horizontal, calc_widget)

        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_layout.setContentsMargins(20, 20, 20, 20)
        left_layout.setSpacing(15)

        self.type_combo = QComboBox()
        self.type_combo.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.type_combo.setToolTip("Select the type of steel section to calculate")
        for s in STEEL_TYPES:
            self.type_combo.addItem(s.name, s.key)

        clear_btn = QPushButton("Clear")
        clear_btn.setObjectName("ClearButton")
        clear_btn.setToolTip("Clear all input fields and reset to default values")

        type_header_layout = QHBoxLayout()
        type_header_layout.addWidget(self.type_combo)
        type_header_layout.addWidget(clear_btn)

        type_box = QGroupBox("1. Select Steel Type")
        type_vbox = QVBoxLayout(type_box)
        type_vbox.addLayout(type_header_layout)

        self.form_host = QWidget()
        self.form_layout = QFormLayout(self.form_host)
        self.form_layout.setLabelAlignment(Qt.AlignLeft)
        self.form_layout.setFormAlignment(Qt.AlignTop)
        self.form_layout.setVerticalSpacing(12)
        self.form_layout.setHorizontalSpacing(20)

        self.inputs: List[QLineEdit] = []
        self.unit_combos: Dict[str, QComboBox] = {}  # Store unit combos for each field
        self.image_box = ImageBox()

        input_box = QGroupBox("2. Section Dimensions")
        input_box.setToolTip("Enter the geometric dimensions of the steel section")
        input_vbox = QVBoxLayout(input_box)
        input_vbox.addWidget(self.form_host)

        self.result_card = QWidget()
        self.result_card.setObjectName("ResultCard")
        card_layout = QVBoxLayout(self.result_card)

        self.result_title = QLabel("CALCULATED THEORETICAL WEIGHT")
        self.result_title.setObjectName("ResultTitle")
        self.result_title.setAlignment(Qt.AlignCenter)

        self.result_label = QLabel("0.00 kg/m")
        self.result_label.setObjectName("ResultLabel")
        self.result_label.setAlignment(Qt.AlignCenter)
        self.result_label.setTextInteractionFlags(Qt.TextSelectableByMouse)

        self.calc_copy_btn = QPushButton("Copy Value")
        self.calc_copy_btn.setObjectName("CalcCopyButton")
        self.calc_copy_btn.setToolTip("Copy the calculated weight to clipboard")
        self.calc_copy_btn.clicked.connect(self.copy_calc_result)

        card_layout.addWidget(self.result_title)
        card_layout.addWidget(self.result_label)
        card_layout.addWidget(self.calc_copy_btn, alignment=Qt.AlignCenter)

        left_layout.addWidget(type_box)
        left_layout.addWidget(input_box)
        left_layout.addWidget(self.result_card)
        left_layout.addStretch(1)

        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(20, 20, 20, 20)
        image_box_wrapper = QGroupBox("Cross-Section Dimension Reference")
        image_box_wrapper.setToolTip(
            "Technical drawing showing dimension labels for the selected steel type"
        )
        img_vbox = QVBoxLayout(image_box_wrapper)
        img_vbox.addWidget(self.image_box)
        right_layout.addWidget(image_box_wrapper)

        main_splitter.addWidget(left_widget)
        main_splitter.addWidget(right_widget)
        main_splitter.setStretchFactor(0, 6)
        main_splitter.setStretchFactor(1, 4)

        calc_layout = QVBoxLayout(calc_widget)
        calc_layout.setContentsMargins(0, 0, 0, 0)
        calc_layout.addWidget(main_splitter)

        self.type_combo.currentIndexChanged.connect(self.rebuild_inputs)
        clear_btn.clicked.connect(self.clear_inputs)
        self.rebuild_inputs()

        self.tabs.addTab(calc_widget, "Manual Calculator")
        logger.info("Calculator tab initialized")

    def rebuild_inputs(self):
        while self.form_layout.count():
            item = self.form_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()
        self.inputs.clear()
        self.unit_combos.clear()

        steel = self.current_type()
        positive_validator = QDoubleValidator(0.001, 999999.0, 3, self)
        positive_validator.setNotation(QDoubleValidator.StandardNotation)

        for label, unit in steel.fields:
            # Create container widget for input + unit combo
            container = QWidget()
            container_layout = QHBoxLayout(container)
            container_layout.setContentsMargins(0, 0, 0, 0)
            container_layout.setSpacing(5)

            edit = QLineEdit()
            edit.setProperty("field_label", label)
            edit.setProperty("original_unit", unit)
            edit.setPlaceholderText("0.00")
            edit.setValidator(positive_validator)
            edit.setToolTip(f"Enter {label} value")

            # Unit combo box
            unit_combo = QComboBox()
            unit_combo.addItems(["mm", "cm", "m", "inch"])
            unit_combo.setCurrentText(
                unit if unit in ["mm", "cm", "m", "inch"] else "mm"
            )
            unit_combo.setMaximumWidth(80)
            unit_combo.setToolTip("Select unit of measurement")
            unit_combo.currentTextChanged.connect(
                lambda text, lbl=label: self.on_unit_changed(lbl, text)
            )

            container_layout.addWidget(edit)
            container_layout.addWidget(unit_combo)

            # Store reference
            self.inputs.append(edit)
            self.unit_combos[label] = unit_combo

            # Label
            label_text = f"<b>{label}</b>"

            if label == "Length" and unit == "m":
                edit.setText("1")
                edit.setToolTip("Length in meters (always in m for beams)")

            edit.textChanged.connect(self.calculate)
            self.form_layout.addRow(label_text, container)

        # Corner Radius field (r1) - optional, for selected steel types
        if steel.key in ["ih", "channel", "angle", "rhs_shs", "tsection"]:
            radius_container = QWidget()
            radius_container_layout = QHBoxLayout(radius_container)
            radius_container_layout.setContentsMargins(0, 0, 0, 0)
            radius_container_layout.setSpacing(5)

            radius_edit = QLineEdit()
            radius_edit.setProperty("field_label", "r1")
            radius_edit.setProperty("original_unit", "mm")
            radius_edit.setText("0")
            radius_edit.setValidator(positive_validator)
            radius_edit.setToolTip("Corner radius (r1). Default 0 for sharp corners, increase for rounded corners")
            radius_edit.textChanged.connect(self.calculate)

            radius_unit_combo = QComboBox()
            radius_unit_combo.addItems(["mm", "cm", "inch"])
            radius_unit_combo.setCurrentText("mm")
            radius_unit_combo.setMaximumWidth(80)
            radius_unit_combo.setToolTip("Select unit for r1")
            radius_unit_combo.currentTextChanged.connect(
                lambda text, lbl="r1": self.on_unit_changed(lbl, text)
            )

            radius_container_layout.addWidget(radius_edit)
            radius_container_layout.addWidget(radius_unit_combo)

            self.inputs.append(radius_edit)
            self.unit_combos["r1"] = radius_unit_combo

            self.form_layout.addRow("<b>r1 (Corner Radius)</b>", radius_container)

        # Quantity field (no unit conversion)
        qty_edit = QLineEdit()
        qty_edit.setProperty("field_label", "Quantity")
        qty_edit.setPlaceholderText("1")
        qty_edit.setText("1")
        qty_edit.setValidator(positive_validator)
        qty_edit.setToolTip("Quantity of steel sections")
        qty_edit.textChanged.connect(self.calculate)
        self.form_layout.addRow("<b>Quantity</b>", qty_edit)
        self.inputs.append(qty_edit)

        image_path = os.path.join(PNG_DIR, steel.image_file)
        self.image_box.set_image(image_path)

        # Set tooltip for the steel type
        self.type_combo.setToolTip(steel.tooltip)

        self.calculate()
        logger.info(f"Input fields rebuilt for steel type: {steel.name}")

    def current_type(self) -> SteelType:
        key = self.type_combo.currentData()
        return next(s for s in STEEL_TYPES if s.key == key)

    def on_unit_changed(self, field_label: str, new_unit: str):
        """Handle unit conversion when user changes unit dropdown"""
        try:
            # Find the input field for this label
            edit = None
            for input_widget in self.inputs:
                if input_widget.property("field_label") == field_label:
                    edit = input_widget
                    break

            if not edit or not edit.text().strip():
                return

            # Get current value and old unit
            old_value = float(edit.text().strip())
            old_unit = edit.property("original_unit")

            # If original unit is meters, don't convert
            if old_unit == "m":
                return

            # Convert from old unit to mm, then to new unit
            old_unit_factor = UNIT_CONVERSION.get(old_unit, 1.0)
            new_unit_factor = UNIT_CONVERSION.get(new_unit, 1.0)

            # Convert to mm first, then to new unit
            value_in_mm = old_value * old_unit_factor
            new_value = value_in_mm / new_unit_factor

            # Update the field
            edit.setText(f"{new_value:.3f}")

            # Update the stored unit
            edit.setProperty("original_unit", new_unit)

            logger.info(
                f"Unit converted for {field_label}: {old_value} {old_unit} -> {new_value:.3f} {new_unit}"
            )

        except Exception as e:
            logger.error(f"Error converting units: {e}")

    def calculate(self):
        steel = self.current_type()
        try:
            values = {}
            for edit in self.inputs:
                label = edit.property("field_label")
                text = edit.text().strip()
                if not text:
                    raise ValueError(f"Missing value for {label}")
                values[label] = float(text)

            # Validate with detailed error messages
            try:
                steel.validator(values)
            except ValueError:
                error_msg = self.get_validation_error_message(steel.key, values)
                self.result_label.setText("Invalid Input")
                self.result_label.setStyleSheet("color: #ef4444;")
                self.current_raw_weight = None
                logger.warning(f"Validation failed for {steel.name}: {error_msg}")
                return

            qty = values.get("Quantity", 1.0)
            base_weight = steel.calc(values)
            total_weight = base_weight * qty

            self.current_raw_weight = total_weight

            if steel.key == "plate":
                suffix = "kg"
            else:
                length_m = values.get("Length", 1.0)
                if abs(length_m - 1.0) < 1e-9 and abs(qty - 1.0) < 1e-9:
                    suffix = "kg/m"
                else:
                    suffix = "kg"

            self.result_label.setText(f"{total_weight:,.2f} {suffix}")
            self.result_label.setStyleSheet("color: #0284c7;")
            logger.info(f"Calculation successful: {total_weight:.2f} {suffix}")
        except (ValueError, KeyError) as e:
            self.result_label.setText("---")
            self.result_label.setStyleSheet("color: #94a3b8;")
            self.current_raw_weight = None
            if str(e):
                logger.warning(f"Calculation error: {e}")

    def get_validation_error_message(self, steel_key: str, values: dict) -> str:
        """Generate user-friendly validation error messages"""
        if steel_key == "plate":
            if values.get("Length", 0) <= 0:
                return "Length must be greater than 0"
            if values.get("Width", 0) <= 0:
                return "Width must be greater than 0"
            if values.get("Thickness", 0) <= 0:
                return "Thickness must be greater than 0"
        elif steel_key in ["ih", "channel"]:
            if values.get("Tw", 0) >= values.get("B", 0):
                return "Web thickness (Tw) must be less than flange width (B)"
            if (2 * values.get("Tf", 0)) >= values.get("H", 0):
                return "Total flange thickness (2×Tf) must be less than height (H)"
        elif steel_key == "tsection":
            if values.get("Tw", 0) >= values.get("B", 0):
                return "Web thickness (Tw) must be less than flange width (B)"
            if values.get("Tf", 0) >= values.get("H", 0):
                return "Flange thickness (Tf) must be less than height (H)"
        elif steel_key == "angle":
            if values.get("Thickness", 0) >= values.get("Leg A", 0):
                return "Thickness must be less than Leg A"
            if values.get("Thickness", 0) >= values.get("Leg B", 0):
                return "Thickness must be less than Leg B"
        elif steel_key == "rhs_shs":
            if (2 * values.get("Thickness", 0)) >= values.get("Width", 0):
                return "Double thickness must be less than Width"
            if (2 * values.get("Thickness", 0)) >= values.get("Height", 0):
                return "Double thickness must be less than Height"
        elif steel_key == "chs":
            if (2 * values.get("Thickness", 0)) >= values.get("OD", 0):
                return "Double thickness must be less than outer diameter (OD)"
        elif steel_key == "rod":
            if values.get("Diameter", 0) <= 0:
                return "Diameter must be greater than 0"
        return "Invalid input values"

    def copy_calc_result(self):
        if self.current_raw_weight is not None:
            QApplication.clipboard().setText(f"{self.current_raw_weight:.2f}")
            logger.info(f"Copied result to clipboard: {self.current_raw_weight:.2f}")

    def clear_inputs(self):
        self.type_combo.blockSignals(True)
        steel = self.current_type()
        for edit in self.inputs:
            lbl = edit.property("field_label")
            if lbl == "Quantity":
                edit.setText("1")
            elif lbl == "Length" and steel.key != "plate":
                edit.setText("1")
            else:
                edit.clear()
        self.type_combo.blockSignals(False)
        self.current_raw_weight = None
        self.calculate()
        logger.info("Input fields cleared")

    # ==========================================
    # LOGIC TAB 2: TRA CỨU ĐA SHEET ĐỘNG (EXCEL)
    # ==========================================
    def init_lookup_tab(self):
        lookup_widget = QWidget()
        splitter = QSplitter(Qt.Horizontal, lookup_widget)

        left_panel = QWidget()
        left_vbox = QVBoxLayout(left_panel)
        left_vbox.setContentsMargins(20, 20, 20, 20)
        left_vbox.setSpacing(10)

        search_box = QGroupBox("Search Section Name")
        search_layout = QVBoxLayout(search_box)
        search_layout.setSpacing(8)

        # Thêm Thanh chọn loại Thư viện/Sheet thép động
        self.lookup_type_combo = QComboBox()
        self.lookup_type_combo.addItem("I Beam / H Beam", "ih")
        self.lookup_type_combo.addItem("PFC / U Channel", "channel")
        self.lookup_type_combo.addItem("Shape VN (HINH_VN)", "hinh_vn")
        self.lookup_type_combo.addItem("Pipe / Tube (Ong,Hop)", "ong_hop")
        self.lookup_type_combo.currentIndexChanged.connect(self.on_lookup_type_changed)

        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("e.g. 'WC 500' or '500 WC'...")

        search_layout.addWidget(QLabel("<b>Select Steel Library:</b>"))
        search_layout.addWidget(self.lookup_type_combo)
        search_layout.addWidget(QLabel("<b>Enter Search Keywords:</b>"))
        search_layout.addWidget(self.search_edit)

        list_box = QGroupBox("Matching Standard Profiles")
        list_layout = QVBoxLayout(list_box)
        self.lookup_list = QListWidget()
        list_layout.addWidget(self.lookup_list)

        left_vbox.addWidget(search_box)
        left_vbox.addWidget(list_box)

        right_panel = QWidget()
        right_vbox = QVBoxLayout(right_panel)
        right_vbox.setContentsMargins(20, 20, 20, 20)

        detail_box = QGroupBox("Profile Specifications Data")
        detail_layout = QVBoxLayout(detail_box)

        self.lookup_card = QWidget()
        self.lookup_card.setObjectName("LookupCard")
        self.card_vbox = QVBoxLayout(self.lookup_card)
        self.card_vbox.setSpacing(10)

        self.reset_lookup_card()  # Trạng thái mặc định ban đầu chưa chọn gì

        detail_layout.addWidget(self.lookup_card)
        right_vbox.addWidget(detail_box)

        splitter.addWidget(left_panel)
        splitter.addWidget(right_panel)
        splitter.setStretchFactor(0, 4)
        splitter.setStretchFactor(1, 6)

        total_layout = QVBoxLayout(lookup_widget)
        total_layout.setContentsMargins(0, 0, 0, 0)
        total_layout.addWidget(splitter)

        self.search_edit.textChanged.connect(self.filter_lookup)
        self.lookup_list.itemClicked.connect(self.display_lookup_details)

        self.filter_lookup()
        self.tabs.addTab(lookup_widget, "Steel Section Lookup")

    def load_data(self):
        """Load data from JSON if available, otherwise from Excel"""
        # Try to load from JSON first
        if os.path.exists(JSON_PATH):
            try:
                logger.info(f"Attempting to load from JSON: {JSON_PATH}")
                import json

                with open(JSON_PATH, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.excel_data = data.get("profiles", [])
                    total_records = len(self.excel_data)
                    logger.info(
                        f"Successfully loaded {total_records} records from JSON"
                    )
                    return
            except Exception as e:
                logger.warning(f"Failed to load JSON: {e}, falling back to Excel")
                self.excel_data = []

        # Fall back to Excel if JSON doesn't exist or failed to load
        self.load_excel_data()

        # Save to JSON for next time
        if self.excel_data:
            self.save_to_json()

    def save_to_json(self):
        """Save loaded data to JSON file for faster loading next time"""
        try:
            logger.info(f"Saving data to JSON: {JSON_PATH}")
            import json

            data = {
                "profiles": self.excel_data,
                "metadata": {
                    "total_records": len(self.excel_data),
                    "source": EXCEL_PATH,
                    "saved_at": "2024-07-16",
                },
            }
            with open(JSON_PATH, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info(f"Successfully saved {len(self.excel_data)} records to JSON")
        except Exception as e:
            logger.error(f"Failed to save JSON: {e}", exc_info=True)

    def load_excel_data(self):
        if not os.path.exists(EXCEL_PATH):
            logger.warning(f"Excel file not found: {EXCEL_PATH}")
            QMessageBox.warning(
                self,
                "File Not Found",
                f"Excel file not found:\n{EXCEL_PATH}\n\nLookup feature will be disabled.",
            )
            return

        try:
            logger.info(f"Loading Excel file: {EXCEL_PATH}")
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

            # Validate required sheets
            required_sheets = ["I lib", "U lib", "HINH_VN", "Ong,Hop"]
            missing_sheets = [s for s in required_sheets if s not in wb.sheetnames]
            if missing_sheets:
                logger.warning(f"Missing sheets: {missing_sheets}")

            def clean_weight_value(val):
                if val is None or str(val).strip() == "" or str(val).strip() == "---":
                    return "---"
                try:
                    return f"{float(val):.2f}"
                except ValueError:
                    return str(val).strip()

            # Nhóm 1: Nạp 2 sheet "I lib" và "U lib" (5 cột D, E, F, N, O)
            type1_sheets = {"ih": "I lib", "channel": "U lib"}
            for key, sname in type1_sheets.items():
                if sname not in wb.sheetnames:
                    logger.warning(f"Sheet not found: {sname}")
                    continue
                sheet = wb[sname]
                rows_iter = sheet.iter_rows(values_only=True)
                try:
                    next(rows_iter)  # Bỏ qua hàng tiêu đề
                except StopIteration:
                    continue
                count = 0
                for row in rows_iter:
                    if len(row) > 3 and row[3] is not None:
                        self.excel_data.append(
                            {
                                "type": key,
                                "D": str(row[3]).strip(),
                                "E": str(row[4]).strip()
                                if len(row) > 4 and row[4] is not None
                                else "---",
                                "F": clean_weight_value(row[5])
                                if len(row) > 5
                                else "---",
                                "N": str(row[13]).strip()
                                if len(row) > 13 and row[13] is not None
                                else "---",
                                "O": clean_weight_value(row[14])
                                if len(row) > 14
                                else "---",
                            }
                        )
                        count += 1
                logger.info(f"Loaded {count} records from {sname}")

            # Nhóm 2: Nạp 2 sheet mới "HINH_VN" và "Ong,Hop" (3 cột A, B, C)
            type2_sheets = {"hinh_vn": "HINH_VN", "ong_hop": "Ong,Hop"}
            for key, sname in type2_sheets.items():
                if sname not in wb.sheetnames:
                    logger.warning(f"Sheet not found: {sname}")
                    continue
                sheet = wb[sname]
                rows_iter = sheet.iter_rows(values_only=True)
                try:
                    next(rows_iter)  # Bỏ qua hàng tiêu đề
                except StopIteration:
                    continue
                count = 0
                for row in rows_iter:
                    if len(row) > 0 and row[0] is not None:
                        self.excel_data.append(
                            {
                                "type": key,
                                "D": str(
                                    row[0]
                                ).strip(),  # Đồng bộ dùng key 'D' để tìm kiếm chung toàn hệ thống
                                "B": clean_weight_value(row[1])
                                if len(row) > 1
                                else "---",
                                "C": str(row[2]).strip()
                                if len(row) > 2 and row[2] is not None
                                else "---",
                            }
                        )
                        count += 1
                logger.info(f"Loaded {count} records from {sname}")

            total_records = len(self.excel_data)
            logger.info(
                f"Excel data loaded successfully: {total_records} total records"
            )
        except Exception as e:
            error_msg = f"Error loading Excel file: {e}"
            logger.error(error_msg, exc_info=True)
            QMessageBox.critical(
                self,
                "Error Loading Excel",
                f"Failed to load Excel file:\n{error_msg}\n\nPlease check the file format and try again.",
            )

    def on_lookup_type_changed(self):
        self.search_edit.clear()
        self.filter_lookup()
        self.reset_lookup_card()

    def reset_lookup_card(self):
        self.clear_layout(self.card_vbox)
        lbl_selected_name = QLabel("No Selection")
        lbl_selected_name.setObjectName("LookupSectionName")
        lbl_selected_name.setAlignment(Qt.AlignCenter)
        self.card_vbox.addWidget(lbl_selected_name)
        self.card_vbox.addStretch(1)

    def clear_layout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clear_layout(item.layout())

    def filter_lookup(self):
        query = self.search_edit.text().lower().strip()
        self.lookup_list.clear()

        if not query:
            # Show placeholder when empty
            item = QListWidgetItem("Enter search keywords to find profiles...")
            item.setFlags(Qt.NoItemFlags)
            self.lookup_list.addItem(item)
            return

        keywords = query.split()
        current_type = self.lookup_type_combo.currentData()

        if not self.excel_data:
            item = QListWidgetItem(
                "⚠️ Excel data not loaded. Check file path or sheet names."
            )
            item.setFlags(Qt.NoItemFlags)
            self.lookup_list.addItem(item)
            return

        matches = 0
        for row in self.excel_data:
            if row["type"] != current_type:
                continue
            name_lower = row["D"].lower()
            if all(kw in name_lower for kw in keywords):
                item = QListWidgetItem(row["D"])
                item.setData(Qt.UserRole, row)
                self.lookup_list.addItem(item)
                matches += 1

        if matches == 0:
            item = QListWidgetItem("No matching profiles found")
            item.setFlags(Qt.NoItemFlags)
            self.lookup_list.addItem(item)

    def display_lookup_details(self, item):
        self.clear_layout(self.card_vbox)
        row_data = item.data(Qt.UserRole)
        if not row_data:
            return

        # Tạo tiêu đề Tên Thép lớn phía trên
        lbl_selected_name = QLabel(row_data["D"])
        lbl_selected_name.setObjectName("LookupSectionName")
        lbl_selected_name.setAlignment(Qt.AlignCenter)
        lbl_selected_name.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.card_vbox.addWidget(lbl_selected_name)

        # Hàm con vẽ dòng dữ liệu gồm (Tiêu đề + Ô text bôi đen + Nút copy số thuần)
        def add_spec_row(title_text, display_text, raw_copy_value):
            title_lbl = QLabel(f"<b>{title_text}:</b>")
            val_lbl = QLabel(display_text)
            val_lbl.setObjectName("LookupValue")
            val_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)

            copy_btn = QPushButton("Copy")
            copy_btn.setObjectName("RowCopyButton")
            if raw_copy_value and raw_copy_value != "---":
                copy_btn.clicked.connect(
                    lambda checked=False, val=raw_copy_value: (
                        QApplication.clipboard().setText(val)
                    )
                )
            else:
                copy_btn.setEnabled(False)

            h_layout = QHBoxLayout()
            h_layout.addWidget(val_lbl, 1)
            h_layout.addWidget(copy_btn)

            self.card_vbox.addWidget(title_lbl)
            self.card_vbox.addLayout(h_layout)

        t = row_data["type"]
        if t in ["ih", "channel"]:
            # Hiển thị cấu trúc 4 trường thông tin cho I và U
            add_spec_row("Original Section", row_data["E"], row_data["E"])

            f_val = row_data["F"]
            disp_f = f"{f_val} kg/m" if f_val != "---" else "---"
            add_spec_row("Weight per meter", disp_f, f_val)

            add_spec_row("Substitute Section", row_data["N"], row_data["N"])

            o_val = row_data["O"]
            disp_o = f"{o_val} kg/m" if o_val != "---" else "---"
            add_spec_row("Weight per meter", disp_o, o_val)
        else:
            # Tự động chuyển đổi hiển thị cấu trúc 2 trường thông tin cho HINH_VN và Ong,Hop
            b_val = row_data["B"]
            disp_b = f"{b_val} kg/m" if b_val != "---" else "---"
            add_spec_row("Weight per meter", disp_b, b_val)

            add_spec_row("Note", row_data["C"], row_data["C"])

        self.card_vbox.addStretch(1)

    # ==========================================
    # QUẢN LÝ GIAO DIỆN (STYLESHEET)
    # ==========================================
    def apply_styles(self):
        self.setStyleSheet("""
            QWidget {
                font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif;
                font-size: 11pt;
                color: #1e293b;
                background-color: #f8fafc;
            }
            QTabWidget::pane {
                border: 1px solid #e2e8f0;
                background: #f8fafc;
                border-radius: 8px;
            }
            QTabBar::tab {
                background: #e2e8f0;
                color: #475569;
                padding: 10px 20px;
                border-top-left-radius: 8px;
                border-top-right-radius: 8px;
                font-weight: 500;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                background: #ffffff;
                color: #0f172a;
                border: 1px solid #e2e8f0;
                border-bottom-color: #f8fafc;
                font-weight: 600;
            }
            QGroupBox {
                font-weight: 600;
                font-size: 11.5pt;
                color: #0f172a;
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                margin-top: 5px;
                padding: 15px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 8px;
                background-color: #f8fafc;
            }
            QLineEdit, QComboBox, QListWidget {
                background: #f8fafc;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                padding: 10px 14px;
                color: #334155;
            }
            QListWidget { padding: 5px; background: #ffffff; }
            QListWidget::item {
                padding: 10px 12px;
                border-radius: 6px;
                margin-bottom: 2px;
            }
            QListWidget::item:hover { background: #f1f5f9; color: #0f172a; }
            QListWidget::item:selected { background: #0284c7; color: white; }
            QLineEdit:focus, QComboBox:focus {
                border: 2px solid #3b82f6;
                background: #ffffff;
            }
            QPushButton#ClearButton {
                background: #f1f5f9;
                color: #64748b;
                border: 1px solid #cbd5e1;
                border-radius: 8px;
                padding: 8px 14px;
                font-weight: 500;
            }
            QPushButton#ClearButton:hover { background: #e2e8f0; color: #334155; }
            
            QPushButton#RowCopyButton {
                background: #ffffff;
                color: #0284c7;
                border: 1px solid #cbd5e1;
                border-radius: 6px;
                padding: 6px 14px;
                font-weight: 600;
            }
            QPushButton#RowCopyButton:hover {
                background: #f1f5f9;
                border-color: #0284c7;
            }
            
            QWidget#ResultCard {
                background-color: #e0f2fe;
                border: 1px solid #bae6fd;
                border-radius: 12px;
                padding: 18px;
            }
            QPushButton#CalcCopyButton {
                background: #ffffff;
                color: #0284c7;
                border: 1px solid #bae6fd;
                border-radius: 8px;
                padding: 8px 20px;
                font-weight: 600;
                margin-top: 5px;
            }
            QPushButton#CalcCopyButton:hover {
                background: #f0f9ff;
                border-color: #0284c7;
            }
            QWidget#LookupCard {
                background-color: #f1f5f9;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                padding: 15px;
            }
            QLabel#ResultTitle {
                color: #64748b;
                font-size: 8.5pt;
                font-weight: 700;
                letter-spacing: 0.5px;
            }
            QLabel#ResultLabel {
                font-size: 26pt;
                font-weight: 800;
                color: #0284c7;
            }
            QLabel#LookupSectionName {
                font-size: 18pt;
                font-weight: 800;
                color: #0f172a;
                border-bottom: 2px solid #cbd5e1;
                padding-bottom: 8px;
                margin-bottom: 5px;
            }
            QLabel#LookupValue {
                font-size: 12pt;
                font-weight: 700;
                color: #0369a1;
                background: white;
                border: 1px solid #e2e8f0;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QLabel { background: transparent; }
            QScrollBar:vertical {
                border: none;
                background: #f8fafc;
                width: 10px;
                margin: 15px 0 15px 0;
            }
            QScrollBar::handle:vertical {
                background: #cbd5e1;
                min-height: 20px;
                border-radius: 5px;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                border: none;
                background: none;
                height: 0px;
            }
        """)


if __name__ == "__main__":
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "1"
    app = QApplication(sys.argv)
    win = MainWindow()
    win.show()
    try:
        sys.exit(app.exec())
    except Exception as e:
        logger.error(f"Application crashed: {e}", exc_info=True)
        raise
