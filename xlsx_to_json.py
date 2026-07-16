import json
import os
from pathlib import Path

import openpyxl

# ==================================================
# CONFIG
# ==================================================

BASE_DIR = Path(__file__).resolve().parent

EXCEL_PATH = BASE_DIR / "alias.xlsx"
OUTPUT_JSON = BASE_DIR / "steel_db.json"

DEBUG = True

TYPE1_SHEETS = {
    "ih": "I lib",
    "channel": "U lib",
}

TYPE2_SHEETS = {
    "hinh_vn": "HINH_VN",
    "ong_hop": "Ong,Hop",
}


# ==================================================
# DEBUG
# ==================================================

def log_info(msg):
    print(f"[INFO] {msg}")

def log_debug(msg):
    if DEBUG:
        print(f"[DEBUG] {msg}")

def log_warning(msg):
    print(f"[WARNING] {msg}")

def log_error(msg):
    print(f"[ERROR] {msg}")


# ==================================================
# HELPERS
# ==================================================

def clean_text(value):
    if value is None:
        return ""

    text = str(value).strip()

    if text == "---":
        return ""

    return text


def clean_weight(value):
    if value is None:
        return None

    text = str(value).strip()

    if text in ("", "---"):
        return None

    try:
        return round(float(value), 3)
    except Exception:
        return None


# ==================================================
# PARSERS
# ==================================================

def parse_type1_sheet(sheet):
    """
    I lib
    U lib

    D E F N O
    """

    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if len(row) < 4:
            continue

        name = clean_text(row[3])

        if not name:
            continue

        rows.append({
            "name": name,
            "original_section": clean_text(row[4]) if len(row) > 4 else "",
            "weight": clean_weight(row[5]) if len(row) > 5 else None,
            "substitute_section": clean_text(row[13]) if len(row) > 13 else "",
            "substitute_weight": clean_weight(row[14]) if len(row) > 14 else None,
        })

    return rows


def parse_type2_sheet(sheet):
    """
    HINH_VN
    Ong,Hop

    A B C
    """

    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if len(row) < 1:
            continue

        name = clean_text(row[0])

        if not name:
            continue

        rows.append({
            "name": name,
            "weight": clean_weight(row[1]) if len(row) > 1 else None,
            "note": clean_text(row[2]) if len(row) > 2 else "",
        })

    return rows


# ==================================================
# MAIN
# ==================================================

def build_database():

    if not EXCEL_PATH.exists():
        raise FileNotFoundError(
            f"Excel file not found: {EXCEL_PATH}"
        )

    log_info(f"Loading workbook: {EXCEL_PATH}")

    workbook = openpyxl.load_workbook(
        EXCEL_PATH,
        data_only=True,
    )

    database = {
        "ih": [],
        "channel": [],
        "hinh_vn": [],
        "ong_hop": [],
    }

    # ------------------------------------------
    # TYPE 1
    # ------------------------------------------

    for key, sheet_name in TYPE1_SHEETS.items():

        if sheet_name not in workbook.sheetnames:
            log_warning(f"Sheet not found: {sheet_name}")
            continue

        sheet = workbook[sheet_name]

        log_debug(f"Reading sheet: {sheet_name}")

        records = parse_type1_sheet(sheet)

        database[key] = records

        log_info(
            f"{sheet_name}: {len(records)} records"
        )

    # ------------------------------------------
    # TYPE 2
    # ------------------------------------------

    for key, sheet_name in TYPE2_SHEETS.items():

        if sheet_name not in workbook.sheetnames:
            log_warning(f"Sheet not found: {sheet_name}")
            continue

        sheet = workbook[sheet_name]

        log_debug(f"Reading sheet: {sheet_name}")

        records = parse_type2_sheet(sheet)

        database[key] = records

        log_info(
            f"{sheet_name}: {len(records)} records"
        )

    return database


def save_json(data):

    with open(
        OUTPUT_JSON,
        "w",
        encoding="utf-8"
    ) as fp:

        json.dump(
            data,
            fp,
            ensure_ascii=False,
            indent=2
        )

    log_info(f"JSON saved: {OUTPUT_JSON}")


def main():

    try:

        data = build_database()

        save_json(data)

        total_records = sum(
            len(v)
            for v in data.values()
        )

        log_info(
            f"Completed successfully "
            f"({total_records} records)"
        )

    except Exception as exc:

        log_error(str(exc))

        raise


if __name__ == "__main__":
    main()