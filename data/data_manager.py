"""Quản lý dữ liệu: load/save JSON, load Excel."""

import json
import os

import openpyxl


class DataManager:
    """Load dữ liệu profile thép từ JSON hoặc Excel."""

    def __init__(self, excel_path: str, json_path: str, logger):
        self.excel_path = excel_path
        self.json_path = json_path
        self.logger = logger

    def load_data(self) -> list:
        """Load data từ JSON nếu có, fallback sang Excel."""
        if os.path.exists(self.json_path):
            try:
                self.logger.info(f"Attempting to load from JSON: {self.json_path}")
                with open(self.json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    profiles = data.get("profiles", [])
                    self.logger.info(f"Successfully loaded {len(profiles)} records from JSON")
                    return profiles
            except Exception as e:
                self.logger.warning(f"Failed to load JSON: {e}, falling back to Excel")

        # Fallback to Excel
        profiles = self._load_excel_data()
        if profiles:
            self._save_to_json(profiles)
        return profiles

    def _save_to_json(self, profiles: list):
        """Save loaded data to JSON file."""
        try:
            self.logger.info(f"Saving data to JSON: {self.json_path}")
            data = {
                "profiles": profiles,
                "metadata": {
                    "total_records": len(profiles),
                    "source": self.excel_path,
                    "saved_at": "2024-07-16",
                },
            }
            with open(self.json_path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            self.logger.info(f"Successfully saved {len(profiles)} records to JSON")
        except Exception as e:
            self.logger.error(f"Failed to save JSON: {e}", exc_info=True)

    def _load_excel_data(self) -> list:
        """Load data from Excel file."""
        if not os.path.exists(self.excel_path):
            self.logger.warning(f"Excel file not found: {self.excel_path}")
            return []

        try:
            self.logger.info(f"Loading Excel file: {self.excel_path}")
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            profiles = []

            # Validate required sheets
            required_sheets = ["I lib", "U lib", "HINH_VN", "Ong,Hop"]
            missing = [s for s in required_sheets if s not in wb.sheetnames]
            if missing:
                self.logger.warning(f"Missing sheets: {missing}")

            def clean_weight(val):
                if val is None or str(val).strip() in ("", "---"):
                    return "---"
                try:
                    return f"{float(val):.2f}"
                except ValueError:
                    return str(val).strip()

            def skip_headers(iterator, data_col_idx=3):
                """Skip header rows: skip None values and known header texts."""
                for row in iterator:
                    if len(row) <= data_col_idx:
                        continue
                    val = row[data_col_idx]
                    if val is None:
                        continue
                    val_upper = str(val).strip().upper()
                    # Skip known header texts
                    if val_upper in ("", "WEIGHT", "PROFILE", "SECTION", "NAME"):
                        continue
                    # Re-yield this row as first data row
                    yield row
                    break
                # Yield remaining rows
                yield from iterator

            # Type 1: "I lib" and "U lib" (5 columns D, E, F, N, O)
            type1 = {"ih": "I lib", "channel": "U lib"}
            for key, sname in type1.items():
                if sname not in wb.sheetnames:
                    self.logger.warning(f"Sheet not found: {sname}")
                    continue
                sheet = wb[sname]
                rows = sheet.iter_rows(values_only=True)
                count = 0
                for row in skip_headers(rows, data_col_idx=3):
                    if len(row) > 3 and row[3] is not None:
                        profiles.append({
                            "type": key,
                            "D": str(row[3]).strip(),
                            "E": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "---",
                            "F": clean_weight(row[5]) if len(row) > 5 else "---",
                            "N": str(row[13]).strip() if len(row) > 13 and row[13] is not None else "---",
                            "O": clean_weight(row[14]) if len(row) > 14 else "---",
                        })
                        count += 1
                self.logger.info(f"Loaded {count} records from {sname}")

            # Type 2: "HINH_VN" and "Ong,Hop" (3 columns A, B, C)
            type2 = {"hinh_vn": "HINH_VN", "ong_hop": "Ong,Hop"}
            for key, sname in type2.items():
                if sname not in wb.sheetnames:
                    self.logger.warning(f"Sheet not found: {sname}")
                    continue
                sheet = wb[sname]
                rows = sheet.iter_rows(values_only=True)
                count = 0
                for row in skip_headers(rows, data_col_idx=0):
                    if len(row) > 0 and row[0] is not None:
                        profiles.append({
                            "type": key,
                            "D": str(row[0]).strip(),
                            "B": clean_weight(row[1]) if len(row) > 1 else "---",
                            "C": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "---",
                        })
                        count += 1
                self.logger.info(f"Loaded {count} records from {sname}")

            self.logger.info(f"Excel data loaded: {len(profiles)} total records")
            return profiles

        except Exception as e:
            self.logger.error(f"Error loading Excel file: {e}", exc_info=True)
            return []